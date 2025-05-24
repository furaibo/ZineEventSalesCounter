import textwrap
import unicodedata
import openpyxl
from openpyxl.styles.borders import Border, Side
from pathlib import Path


# テーブル描画
def write_excel_table(
        sheet, sheet_data: list,
        count_empty_row=0, count_empty_col=0):

    # 罫線の設定
    side = Side(style="thin", color="000000")
    border = Border(top=side, bottom=side, left=side, right=side)

    # テーブルの描画
    for row_i, row_data in enumerate(sheet_data):
        for col_i, data in enumerate(row_data):
            # セルの位置(列)の取得
            # Note: 列アルファベット名の生成(A,B,...)および列番号を自動生成
            col_letter = chr(ord("A") + col_i + count_empty_col)
            address = f"{col_letter}{row_i + 1 + count_empty_row}"

            # 値の格納と罫線の設定
            sheet[address] = data
            sheet[address].border = border


# 入力内容に応じた列幅の調整
def adjust_table_columns(sheet):
    for i, col in enumerate(sheet.columns):
        max_length = 0

        # マルチバイト文字に対応したセル文字内列幅の判定
        # Note: マルチバイト文字は半角2文字分の幅として判定
        for cell in col:
            cell_length = 0
            for c in str(cell.value):
                if unicodedata.east_asian_width(c) in "FWA":
                    cell_length += 2
                else:
                    cell_length += 1
            if cell_length > max_length:
                max_length = cell_length

        # セルの位置(列)の取得
        # Note: 列アルファベット名の生成(A,B,...)
        col_letter = chr(ord("A") + i)

        # 余裕をもたせた幅の値を設定(最長の文字列長の1.1倍)
        adjusted_width = max_length * 1.1
        sheet.column_dimensions[col_letter].width = adjusted_width


# Excel形式レポートの作成処理
def generate_excel_report(save_path: Path, data_dict: dict):
    # workbookの作成
    wb = openpyxl.Workbook()

    # 1枚目シート作成
    # Note: デフォルトの「Sheet」を参照、名称変更
    sheet1 = wb["Sheet"]
    sheet1.title = "イベント情報"

    # イベント情報作成
    sheet1_data = []
    sheet1_data.extend([
        ["イベント名", data_dict["zine_event"]["name"]],
        ["実施日時", data_dict["zine_event"]["date"]],
        ["内容説明", "/n".join(
            textwrap.wrap(data_dict["zine_event"]["desc"], 40))],
    ])
    write_excel_table(sheet1, sheet1_data, 1, 1)
    adjust_table_columns(sheet1)

    # 2枚目シート作成
    sheet2 = wb.create_sheet(title="頒布数集計結果")

    # 頒布数情報作成
    sheet2_data = []
    sheet2_data.append(["タイトル", "頒布数", "集計金額"])
    for item in data_dict["sales_tally"]:
        sheet2_data.append([
            item["product"],
            item["sales_count"],
            item["sales_amount"]
        ])
    write_excel_table(sheet2, sheet2_data, 1, 1)
    adjust_table_columns(sheet2)

    # 3枚目シート作成
    sheet3 = wb.create_sheet(title="頒布履歴情報")

    # 頒布履歴情報作成
    sheet3_data = []
    sheet3_data.append(["タイトル", "頒布価格", "頒布日時"])
    for item in data_dict["sales_history"]:
        sheet3_data.append([
            item["product"],
            item["price"],
            item["sales_at"],
        ])
    write_excel_table(sheet3, sheet3_data, 1, 1)
    adjust_table_columns(sheet3)

    # ファイルの保存
    wb.save(save_path.as_posix())
